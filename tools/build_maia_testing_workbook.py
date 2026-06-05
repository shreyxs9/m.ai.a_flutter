from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path("outputs") / "019e8d83-a753-77b3-8d19-e15ec2daaaae"
OUTPUT_FILE = OUTPUT_DIR / "maia_flutter_testing_tracker.xlsx"


STATUS_VALUES = ["Not Started", "In Progress", "Blocked", "Passed", "Failed", "Retest", "N/A"]
PRIORITY_VALUES = ["P0", "P1", "P2", "P3"]
SEVERITY_VALUES = ["Critical", "High", "Medium", "Low"]
ENV_VALUES = ["Flutter Web Chrome", "Flutter Web Edge", "Mobile Web Android", "Mobile Web iOS", "Android APK", "Backend API", "Firebase Hosting"]


def d(day: int) -> date:
    return date(2026, 6, day)


timeline_rows = [
    [d(3), "Kickoff and baseline", "Freeze testing scope, confirm environments, run flutter analyze/test/build web, capture current blockers", "QA Lead", "In Progress", "Build passes, environments confirmed, sheet ownership assigned"],
    [d(4), "Auth and invite flows", "Google OAuth, dev login, tenant switch, workspace invite, project invite, session persistence", "QA", "Not Started", "All auth routes pass on Chrome and Edge"],
    [d(5), "Projects and settings", "Project list, create/join, project settings, members, roles, check-in time/timezone, archive/delete", "QA", "Not Started", "Project lifecycle verified with admin/member users"],
    [d(6), "Chat, AI, media, search", "Project chat, mentions, replies, relays, broadcast, blockers, media download, poll status, search", "QA + Dev", "Not Started", "Chat flows stable and no stuck pending inference"],
    [d(8), "Profile and Google Sheets", "Profile account, theme/timezone/title, Google Sheets connect/disconnect/attach/detach", "QA", "Not Started", "Connector and profile behavior match release expectations"],
    [d(9), "Admin console", "Members/projects tabs, role changes, remove member, archive/unarchive, copy invite, audit placeholder", "QA + Product", "Not Started", "Admin permissions and destructive warnings verified"],
    [d(10), "PWA, push, mobile web", "Manifest, service worker scope, push prompt, Firebase config, Android/iOS mobile web, safe areas", "QA", "Not Started", "PWA/push limitations documented and mobile web approved"],
    [d(11), "Integration regression", "Backend API contracts, tenant headers, API_ORIGIN, OAuth callback, hosting rewrite assumptions", "Dev + QA", "Not Started", "No critical integration defects open"],
    [d(12), "Fix sprint 1", "Retest P0/P1 issues, stabilize release branch, verify logs and error states", "Dev", "Not Started", "P0 fixed; P1 triaged with owners"],
    [d(15), "Full regression", "Run all P0/P1/P2 tests across target environments", "QA", "Not Started", "All critical paths passed or waived"],
    [d(16), "UAT prep", "Prepare demo accounts, test data, release notes, support runbook, client UAT script", "Product + QA", "Not Started", "UAT pack ready for client"],
    [d(17), "Client UAT", "Run client-facing acceptance script and capture issues/signoff comments", "Client + Product", "Not Started", "Client issues logged with severity"],
    [d(18), "Release readiness", "Final release gates, backup/rollback, production env, Firebase Hosting, OAuth redirect allowlist", "DevOps + QA", "Not Started", "Go/no-go checklist complete"],
    [d(19), "Client handoff", "Deliver workbook, release notes, known limitations, support contacts, client signoff", "Product", "Not Started", "Client handoff complete"],
]


test_cases = [
    ("AUTH-001", "Auth", "Open unauthenticated /", "User is redirected to /login without errors", "P0", "Critical", "Flutter Web Chrome"),
    ("AUTH-002", "Auth", "Open /welcome unauthenticated", "Login screen renders and does not loop", "P1", "High", "Flutter Web Chrome"),
    ("AUTH-003", "Auth", "Complete Google OAuth callback with token", "Session persists and app lands on projects", "P0", "Critical", "Flutter Web Chrome"),
    ("AUTH-004", "Auth", "Reload after login", "Session and selected tenant are retained", "P0", "Critical", "Flutter Web Chrome"),
    ("AUTH-005", "Auth", "Logout or expired token path", "User returns to login and protected data is cleared", "P0", "Critical", "Flutter Web Chrome"),
    ("AUTH-006", "Auth", "Use dev-login in local mode", "Available users can be selected and session is created", "P2", "Medium", "Backend API"),
    ("INV-001", "Invites", "Open /join/{projectCode} while signed out", "Invite is saved, login starts, and project join resumes", "P0", "Critical", "Flutter Web Chrome"),
    ("INV-002", "Invites", "Open /join-workspace/{code} while signed out", "Workspace preview/join flow survives auth redirect", "P0", "Critical", "Flutter Web Chrome"),
    ("INV-003", "Invites", "Enter lowercase project invite code", "Code is trimmed, uppercased, and matched", "P1", "High", "Flutter Web Chrome"),
    ("INV-004", "Invites", "Join project already joined", "Clear already-member message and navigation to project", "P2", "Medium", "Flutter Web Chrome"),
    ("DASH-001", "Projects", "Load projects dashboard", "Project cards, blocker totals, counts, and empty state render", "P0", "Critical", "Flutter Web Chrome"),
    ("DASH-002", "Projects", "Refresh project list", "List reloads without duplicate cards or stale blockers", "P1", "High", "Flutter Web Chrome"),
    ("DASH-003", "Projects", "Create project with name only", "Project is created and user navigates to project chat", "P0", "Critical", "Flutter Web Chrome"),
    ("DASH-004", "Projects", "Create project with description and icon", "Metadata appears in list and settings", "P1", "High", "Flutter Web Chrome"),
    ("DASH-005", "Projects", "Open empty dashboard as new user", "No projects state offers new/join project actions", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-001", "Chat", "Open /project/{id}", "Thread loads, project header and context panel are visible", "P0", "Critical", "Flutter Web Chrome"),
    ("CHAT-002", "Chat", "Send normal message", "User message appears and API response is handled", "P0", "Critical", "Flutter Web Chrome"),
    ("CHAT-003", "Chat", "Send message that returns pending", "Pending state polls until inference completes or errors cleanly", "P0", "Critical", "Flutter Web Chrome"),
    ("CHAT-004", "Chat", "Reply to message", "Reply metadata is sent and UI shows reply context", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-005", "Chat", "Mention teammate", "Mention payload contains selected user IDs", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-006", "Chat", "Relay message to one teammate", "Relay endpoint is called and returned messages render", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-007", "Chat", "Broadcast to project", "Broadcast endpoint is called and confirmations render", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-008", "Chat", "Resolve and unresolve blocker", "Message state toggles and blocker count updates", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-009", "Chat", "Search project messages", "Search results match query and open correct context", "P1", "High", "Flutter Web Chrome"),
    ("CHAT-010", "Chat", "Open media attachment download URL", "Signed media URL is fetched and usable before expiry", "P2", "Medium", "Flutter Web Chrome"),
    ("CTX-001", "Context", "Open project context panel", "Goals, state, team status, and history sections load", "P1", "High", "Flutter Web Chrome"),
    ("CTX-002", "Context", "Open member timeline", "Timeline loads only selected member activity", "P2", "Medium", "Flutter Web Chrome"),
    ("CTX-003", "Context", "History endpoint fails", "Panel shows useful error and retry path", "P2", "Medium", "Flutter Web Chrome"),
    ("SET-001", "Project Settings", "Open settings sheet", "Current project name/icon/check-in settings populate", "P0", "Critical", "Flutter Web Chrome"),
    ("SET-002", "Project Settings", "Edit name/icon/check-in time/timezone/digest delay", "Patch payload is correct and changes persist after reload", "P0", "Critical", "Flutter Web Chrome"),
    ("SET-003", "Project Settings", "Add project member", "Eligible tenant member is added and list refreshes", "P1", "High", "Flutter Web Chrome"),
    ("SET-004", "Project Settings", "Change project member role", "Role patch succeeds and UI reflects new permission", "P1", "High", "Flutter Web Chrome"),
    ("SET-005", "Project Settings", "Remove project member", "Confirmation appears and member loses access", "P1", "High", "Flutter Web Chrome"),
    ("SET-006", "Project Settings", "Copy project invite code", "Clipboard receives correct invite link/code", "P2", "Medium", "Flutter Web Chrome"),
    ("SET-007", "Project Settings", "Delete project as unauthorized user", "Delete action is blocked with permission explanation", "P0", "Critical", "Flutter Web Chrome"),
    ("SET-008", "Project Settings", "Delete project as authorized user", "Confirmation appears and project disappears for members", "P0", "Critical", "Flutter Web Chrome"),
    ("GS-001", "Google Sheets", "Open Google Sheets status", "Connected/disconnected state is accurate", "P1", "High", "Flutter Web Chrome"),
    ("GS-002", "Google Sheets", "Start Google Sheets connect", "Browser navigates to backend connect URL", "P1", "High", "Flutter Web Chrome"),
    ("GS-003", "Google Sheets", "Disconnect Google Sheets", "Backend disconnect succeeds and UI refreshes", "P1", "High", "Flutter Web Chrome"),
    ("GS-004", "Google Sheets", "Attach valid sheet to project", "Sheet appears with label and schema hint", "P1", "High", "Flutter Web Chrome"),
    ("GS-005", "Google Sheets", "Detach sheet from project", "Confirmation appears and sheet is removed", "P1", "High", "Flutter Web Chrome"),
    ("GS-006", "Google Sheets", "Attach invalid sheet ID", "Validation/error message is clear", "P2", "Medium", "Flutter Web Chrome"),
    ("PROF-001", "Profile", "Open /profile/account", "Account details load without admin flicker", "P1", "High", "Flutter Web Chrome"),
    ("PROF-002", "Profile", "Update display name/title", "Updated value persists and app state refreshes", "P1", "High", "Flutter Web Chrome"),
    ("PROF-003", "Profile", "Update timezone", "maia_timezone key behavior and backend timezone update are correct", "P1", "High", "Flutter Web Chrome"),
    ("PROF-004", "Profile", "Switch theme", "Theme persists and reload keeps selection", "P2", "Medium", "Flutter Web Chrome"),
    ("PROF-005", "Profile", "Open notifications section", "Notification settings and push state render", "P2", "Medium", "Flutter Web Chrome"),
    ("ADMIN-001", "Admin", "Open /admin as non-admin", "UI blocks access or hides admin entry", "P0", "Critical", "Flutter Web Chrome"),
    ("ADMIN-002", "Admin", "Open /admin as admin", "Members/projects tabs load", "P0", "Critical", "Flutter Web Chrome"),
    ("ADMIN-003", "Admin", "Search/filter/sort members", "Results update correctly and layout remains usable", "P2", "Medium", "Flutter Web Chrome"),
    ("ADMIN-004", "Admin", "Promote/demote tenant member", "Role change succeeds except self-protection cases", "P0", "Critical", "Flutter Web Chrome"),
    ("ADMIN-005", "Admin", "Remove tenant member", "Confirmation appears and member loses workspace access", "P0", "Critical", "Flutter Web Chrome"),
    ("ADMIN-006", "Admin", "Archive/unarchive project", "Project state toggles and list refreshes", "P1", "High", "Flutter Web Chrome"),
    ("ADMIN-007", "Admin", "Copy workspace invite", "Invite link uses active browser origin", "P2", "Medium", "Flutter Web Chrome"),
    ("ADMIN-008", "Admin", "Open audit section", "Placeholder communicates backend dependency without broken UI", "P3", "Low", "Flutter Web Chrome"),
    ("PWA-001", "PWA", "Load web manifest", "Icons, theme color, scope, and start URL are valid", "P1", "High", "Flutter Web Chrome"),
    ("PWA-002", "PWA", "Validate service worker scopes", "Flutter worker and Firebase messaging worker do not collide", "P0", "Critical", "Flutter Web Chrome"),
    ("PWA-003", "PWA", "Push prompt when Firebase config missing", "App logs/configures expected skip without broken UI", "P2", "Medium", "Flutter Web Chrome"),
    ("PWA-004", "PWA", "Push registration with Firebase config", "Token registers through /me/push-tokens", "P1", "High", "Flutter Web Chrome"),
    ("PWA-005", "PWA", "Dismiss push prompt", "maia_push_prompt_dismissed_at persists", "P2", "Medium", "Flutter Web Chrome"),
    ("MOB-001", "Mobile Web", "Open app on Android phone web-server", "Safe area, navigation, sheets, and chat composer fit", "P0", "Critical", "Mobile Web Android"),
    ("MOB-002", "Mobile Web", "Open app on iOS Safari", "Known push limits documented, layout still usable", "P1", "High", "Mobile Web iOS"),
    ("MOB-003", "Mobile Web", "Use keyboard in chat composer", "Composer remains visible and no content overlap", "P1", "High", "Mobile Web Android"),
    ("MOB-004", "Mobile Web", "Open settings bottom sheet on small screen", "Sheet scrolls and buttons remain reachable", "P1", "High", "Mobile Web Android"),
    ("API-001", "Backend API", "Verify Authorization and X-Tenant-Id headers", "All authenticated calls include required headers", "P0", "Critical", "Backend API"),
    ("API-002", "Backend API", "API base URL/path dart-defines", "Local and hosted requests target expected /api/v1 base", "P0", "Critical", "Backend API"),
    ("API-003", "Backend API", "API error response", "ApiException surfaces clear user-facing message", "P1", "High", "Backend API"),
    ("REL-001", "Release", "Run flutter analyze", "No analyzer errors", "P0", "Critical", "Flutter Web Chrome"),
    ("REL-002", "Release", "Run flutter test", "All automated tests pass", "P0", "Critical", "Flutter Web Chrome"),
    ("REL-003", "Release", "Run flutter build web --release", "Release web build completes", "P0", "Critical", "Flutter Web Chrome"),
    ("REL-004", "Release", "Verify Firebase Hosting rewrites", "Deep links and /api/v1 rewrites work", "P0", "Critical", "Firebase Hosting"),
    ("REL-005", "Release", "Verify OAuth redirect allowlist", "Production OAuth callback returns to Flutter origin", "P0", "Critical", "Firebase Hosting"),
    ("REL-006", "Release", "Android APK readiness", "JDK/Gradle/device readiness or documented blocker", "P2", "Medium", "Android APK"),
]


uat_rows = [
    ["UAT-001", "Client can log in with approved Google account", "Client", "Pending", ""],
    ["UAT-002", "Client can see the correct workspace and projects", "Client", "Pending", ""],
    ["UAT-003", "Client can create or join a project using invite flow", "Client", "Pending", ""],
    ["UAT-004", "Client can send project chat messages and receive Maia responses", "Client", "Pending", ""],
    ["UAT-005", "Client can manage project members and settings", "Client Admin", "Pending", ""],
    ["UAT-006", "Client can connect and attach Google Sheets where applicable", "Client Admin", "Pending", ""],
    ["UAT-007", "Client admin can manage workspace members/projects", "Client Admin", "Pending", ""],
    ["UAT-008", "Client accepts known limitations: audit placeholder, push/browser limits, Android APK dependency", "Client", "Pending", ""],
    ["UAT-009", "Client receives release notes, support path, and handoff checklist", "Product", "Pending", ""],
]


release_gates = [
    ["Automated checks", "flutter analyze passes", "P0", "Open", "QA/Dev", "Run before every release candidate"],
    ["Automated checks", "flutter test passes", "P0", "Open", "QA/Dev", "Minimum expected automated model/widget coverage"],
    ["Build", "flutter build web --release passes", "P0", "Open", "Dev", "Use production dart-defines"],
    ["Config", "MAIA_API_BASE_URL, MAIA_API_BASE_PATH, API_ORIGIN confirmed", "P0", "Open", "DevOps", "Local, staging, production values listed"],
    ["Hosting", "Firebase Hosting deep-link and /api/v1 rewrites verified", "P0", "Open", "DevOps", "Test /project/{id}, /profile/account, /join/{code}"],
    ["OAuth", "Google OAuth redirect URLs allow Flutter hosted origin", "P0", "Open", "DevOps", "Production callback tested"],
    ["Push", "Firebase web app values and VAPID key configured or limitation documented", "P1", "Open", "DevOps", "Browser push needs valid config"],
    ["Data", "Demo/test users and projects cleaned or clearly marked", "P1", "Open", "QA", "Avoid leaking test fixtures to client workspace"],
    ["Security", "Non-admin cannot access admin console or destructive actions", "P0", "Open", "QA", "Backend remains final permission authority"],
    ["Client", "UAT signoff captured", "P0", "Open", "Product", "Written signoff before handoff"],
    ["Support", "Known limitations and rollback/support contact shared", "P1", "Open", "Product", "Include support owner and escalation path"],
]


env_rows = [
    ["Local Web", "Chrome", "flutter run -d chrome --web-port 3000", "https://maia-backend-7vtst4xamq-el.a.run.app/api/v1", "API_ORIGIN=https://maia-backend-7vtst4xamq-el.a.run.app", "Primary dev validation"],
    ["Local LAN", "Android phone browser", "flutter run -d web-server --web-hostname 0.0.0.0 --web-port 3000", "http://<LAN-IP>:8000/api/v1", "API_ORIGIN=http://<LAN-IP>:8000", "Phone web validation"],
    ["Staging Web", "Chrome/Edge", "flutter build web --release", "Staging /api/v1 rewrite", "Staging OAuth redirect", "Client preview before production"],
    ["Production Web", "Chrome/Edge/Mobile Safari", "Firebase Hosting deploy", "Production /api/v1 rewrite", "Production OAuth redirect", "Client handoff target"],
    ["Android APK", "Android 13+", "flutter build apk --release", "Production API URL", "Production OAuth redirect", "Optional until JDK/Gradle path is confirmed"],
]


handoff_rows = [
    ["Release notes prepared", "Product", "Open", ""],
    ["Known limitations documented", "Product", "Open", "Audit placeholder, Google picker parity, push/browser limits, Android APK readiness"],
    ["Production URL shared", "DevOps", "Open", ""],
    ["Client admin users confirmed", "Product", "Open", ""],
    ["Support contact and escalation path shared", "Product", "Open", ""],
    ["Rollback/restore plan documented", "DevOps", "Open", ""],
    ["Test tracker reviewed with client", "QA", "Open", ""],
    ["Client UAT signoff received", "Client", "Open", ""],
    ["Go-live approval recorded", "Product", "Open", ""],
]


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:80])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)


def add_table_validations(ws, status_col=None, priority_col=None, severity_col=None, env_col=None, max_row=250):
    def add_list(col, values):
        dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")

    if status_col:
        add_list(status_col, STATUS_VALUES)
    if priority_col:
        add_list(priority_col, PRIORITY_VALUES)
    if severity_col:
        add_list(severity_col, SEVERITY_VALUES)
    if env_col:
        add_list(env_col, ENV_VALUES)


def add_status_rules(ws, status_col, max_row=250):
    fills = {
        "Passed": "C6EFCE",
        "Failed": "FFC7CE",
        "Blocked": "F4B084",
        "In Progress": "BDD7EE",
        "Retest": "FFE699",
    }
    for status, color in fills.items():
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{max_row}",
            FormulaRule(formula=[f'${status_col}2="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )


def build():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    dash = wb.create_sheet("Dashboard")
    timeline = wb.create_sheet("Timeline")
    cases = wb.create_sheet("Test Cases")
    defects = wb.create_sheet("Defect Log")
    uat = wb.create_sheet("UAT Signoff")
    gates = wb.create_sheet("Release Gates")
    env = wb.create_sheet("Environment Matrix")
    handoff = wb.create_sheet("Client Handoff")
    data = wb.create_sheet("Test Data")

    timeline.append(["Date", "Phase", "Scope", "Owner", "Status", "Exit Criteria"])
    for row in timeline_rows:
        timeline.append(row)
    add_table_validations(timeline, status_col="E", max_row=100)
    add_status_rules(timeline, "E", max_row=100)
    style_sheet(timeline)
    for cell in timeline["A"]:
        if cell.row > 1:
            cell.number_format = "yyyy-mm-dd"

    cases.append(["Test ID", "Area", "Scenario", "Expected Result", "Priority", "Severity", "Environment", "Owner", "Status", "Run Date", "Evidence / Notes", "Defect ID"])
    for case in test_cases:
        cases.append([*case, "", "Not Started", "", "", ""])
    add_table_validations(cases, status_col="I", priority_col="E", severity_col="F", env_col="G", max_row=300)
    add_status_rules(cases, "I", max_row=300)
    style_sheet(cases)
    for cell in cases["J"]:
        if cell.row > 1:
            cell.number_format = "yyyy-mm-dd"

    defects.append(["Defect ID", "Found Date", "Area", "Summary", "Severity", "Priority", "Environment", "Owner", "Status", "Linked Test ID", "Fix Due", "Resolution Notes"])
    for i in range(1, 31):
        defects.append([f"BUG-{i:03d}", "", "", "", "", "", "", "", "Open" if i == 1 else "", "", "", ""])
    add_table_validations(defects, status_col="I", priority_col="F", severity_col="E", env_col="G", max_row=120)
    add_status_rules(defects, "I", max_row=120)
    style_sheet(defects)
    for col in ("B", "K"):
        for cell in defects[col]:
            if cell.row > 1:
                cell.number_format = "yyyy-mm-dd"

    uat.append(["UAT ID", "Acceptance Item", "Owner", "Status", "Client Comments"])
    for row in uat_rows:
        uat.append(row)
    add_table_validations(uat, status_col="D", max_row=100)
    add_status_rules(uat, "D", max_row=100)
    style_sheet(uat)

    gates.append(["Gate Category", "Gate", "Priority", "Status", "Owner", "Notes"])
    for row in release_gates:
        gates.append(row)
    add_table_validations(gates, status_col="D", priority_col="C", max_row=100)
    add_status_rules(gates, "D", max_row=100)
    style_sheet(gates)

    env.append(["Environment", "Target", "Run Command / Deployment", "API Base", "Auth / Redirect Config", "Purpose"])
    for row in env_rows:
        env.append(row)
    style_sheet(env)

    handoff.append(["Checklist Item", "Owner", "Status", "Notes"])
    for row in handoff_rows:
        handoff.append(row)
    add_table_validations(handoff, status_col="C", max_row=100)
    add_status_rules(handoff, "C", max_row=100)
    style_sheet(handoff)

    data.append(["Data Set", "Purpose", "Required Records", "Status", "Notes"])
    for row in [
        ["Admin user", "Admin console and destructive-permission tests", "Tenant admin and super_admin", "Needed", ""],
        ["Member user", "Normal chat/project/profile tests", "Standard workspace member", "Needed", ""],
        ["Non-member user", "Invite and denied-access tests", "Google account outside tenant", "Needed", ""],
        ["Project with members", "Chat/settings/context tests", "At least 3 members with mixed roles", "Needed", ""],
        ["Project with blockers", "Dashboard/chat blocker count tests", "Open and resolved blocker messages", "Needed", ""],
        ["Google Sheet", "Connector attach/detach tests", "Accessible test sheet ID", "Needed", ""],
        ["Push config", "Browser push tests", "Firebase public config and VAPID key", "Optional", "Document if unavailable"],
    ]:
        data.append(row)
    style_sheet(data)

    dash["A1"] = "M.AI.A Flutter Whole-App Testing Tracker"
    dash["A1"].font = Font(size=18, bold=True, color="1F4E78")
    dash["A2"] = "Plan starts"
    dash["B2"] = d(3)
    dash["B2"].number_format = "yyyy-mm-dd"
    dash["C2"] = "Target client handoff"
    dash["D2"] = d(19)
    dash["D2"].number_format = "yyyy-mm-dd"
    dash["A4"] = "Metric"
    dash["B4"] = "Value"
    dash["A5"] = "Total test cases"
    dash["B5"] = "=COUNTA('Test Cases'!A2:A300)"
    dash["A6"] = "Passed"
    dash["B6"] = '=COUNTIF(\'Test Cases\'!I:I,"Passed")'
    dash["A7"] = "Failed"
    dash["B7"] = '=COUNTIF(\'Test Cases\'!I:I,"Failed")'
    dash["A8"] = "Blocked"
    dash["B8"] = '=COUNTIF(\'Test Cases\'!I:I,"Blocked")'
    dash["A9"] = "Not started"
    dash["B9"] = '=COUNTIF(\'Test Cases\'!I:I,"Not Started")'
    dash["A10"] = "Completion %"
    dash["B10"] = '=IF(B5=0,0,B6/B5)'
    dash["B10"].number_format = "0%"
    dash["A11"] = "Open critical defects"
    dash["B11"] = '=COUNTIFS(\'Defect Log\'!E:E,"Critical",\'Defect Log\'!I:I,"<>Closed",\'Defect Log\'!I:I,"<>")'
    dash["A12"] = "Open release gates"
    dash["B12"] = '=COUNTIF(\'Release Gates\'!D:D,"Open")'
    dash["A14"] = "Release decision"
    dash["B14"] = '=IF(AND(B7=0,B8=0,B11=0,B12=0),"Ready for client handoff","Not ready")'
    dash["D4"] = "Status"
    dash["E4"] = "Count"
    for idx, status in enumerate(STATUS_VALUES, start=5):
        dash[f"D{idx}"] = status
        dash[f"E{idx}"] = f'=COUNTIF(\'Test Cases\'!I:I,D{idx})'
    dash["G4"] = "Area"
    dash["H4"] = "Count"
    areas = sorted({case[1] for case in test_cases})
    for idx, area in enumerate(areas, start=5):
        dash[f"G{idx}"] = area
        dash[f"H{idx}"] = f'=COUNTIF(\'Test Cases\'!B:B,G{idx})'

    style_sheet(dash, freeze="A4")
    for row in range(1, 20):
        dash.row_dimensions[row].height = 22
    for col, width in {"A": 28, "B": 18, "C": 24, "D": 18, "E": 12, "G": 20, "H": 12}.items():
        dash.column_dimensions[col].width = width

    bar = BarChart()
    bar.title = "Test Cases by Status"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Status"
    bar.add_data(Reference(dash, min_col=5, min_row=4, max_row=11), titles_from_data=True)
    bar.set_categories(Reference(dash, min_col=4, min_row=5, max_row=11))
    bar.height = 7
    bar.width = 12
    dash.add_chart(bar, "D14")

    pie = PieChart()
    pie.title = "Coverage by Area"
    pie.add_data(Reference(dash, min_col=8, min_row=4, max_row=4 + len(areas)), titles_from_data=True)
    pie.set_categories(Reference(dash, min_col=7, min_row=5, max_row=4 + len(areas)))
    pie.height = 7
    pie.width = 12
    dash.add_chart(pie, "G14")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)


def verify():
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    required = [
        "Dashboard",
        "Timeline",
        "Test Cases",
        "Defect Log",
        "UAT Signoff",
        "Release Gates",
        "Environment Matrix",
        "Client Handoff",
        "Test Data",
    ]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    if wb["Test Cases"].max_row < 70:
        raise RuntimeError("Expected at least 70 test-case rows")
    first_timeline_date = wb["Timeline"]["A2"].value
    if hasattr(first_timeline_date, "date"):
        first_timeline_date = first_timeline_date.date()
    if first_timeline_date != d(3):
        raise RuntimeError("Timeline does not start on expected date")
    if not str(wb["Dashboard"]["B14"].value).startswith("=IF("):
        raise RuntimeError("Dashboard release-decision formula missing")
    return {
        "sheets": wb.sheetnames,
        "test_case_rows": wb["Test Cases"].max_row - 1,
        "timeline_rows": wb["Timeline"].max_row - 1,
        "path": str(OUTPUT_FILE.resolve()),
    }


if __name__ == "__main__":
    build()
    result = verify()
    print(result)
